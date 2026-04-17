#!/usr/bin/env python3
"""Generate Fraser Health hospitalist invoice workbooks from the master schedule."""

from __future__ import annotations

import argparse
import csv
import html
import io
import os
import re
import shutil
import subprocess
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from email import policy
from email.parser import BytesParser
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Iterable
from urllib.parse import parse_qs, quote, unquote, urlparse
from urllib.request import urlopen
from uuid import uuid4

OPENPYXL_IMPORT_ERROR: ModuleNotFoundError | None = None

try:
    from openpyxl import load_workbook
    from openpyxl.workbook.properties import CalcProperties
except ModuleNotFoundError as exc:
    if exc.name != "openpyxl":
        raise
    OPENPYXL_IMPORT_ERROR = exc
    load_workbook = None
    CalcProperties = None


BASE_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = BASE_DIR / "outputs"
SCRIPT_PATH = BASE_DIR / "hospitalist_invoice_generator.py"
BUNDLED_PYTHON = Path(
    r"C:\Users\bobg6\.cache\codex-runtimes\codex-primary-runtime\dependencies\python\python.exe"
)
DEFAULT_MASTER_URL = (
    "https://docs.google.com/spreadsheets/d/"
    "1DlDnljstmxsqgJyKTU0eWdl4-YBRQvm3eaE7PENVVRE/edit?gid=61161592"
)
DEFAULT_TEMPLATE = Path(
    BASE_DIR / "templates" / "FHA Invoice 2025 Hospitalist All Hours Final - 02172026.xlsx"
)
START_ROW = 18
END_ROW = 217
ADMIN_MINUTES = 15
ON_CALL_DAY_SHIFT_START = time(6, 30)
ON_CALL_DAY_SHIFT_END = time(17, 30)
ADMIN_TYPE_LABELS = {
    "scheduled": "Scheduled daytime/team shifts",
    "evening": "Evening shifts",
    "overnight": "Overnight shifts",
    "admit": "ADMIT shifts",
    "virtual": "VIRTUAL shifts",
}


def ensure_openpyxl_available() -> None:
    if OPENPYXL_IMPORT_ERROR is None:
        return

    bundled_hint = ""
    if BUNDLED_PYTHON.exists():
        bundled_hint = (
            f"\nRecommended launch command:\n"
            f'  "{BUNDLED_PYTHON}" "{SCRIPT_PATH}" --serve\n'
        )

    raise SystemExit(
        "This script needs the Python package 'openpyxl', but it is not installed in the Python you used to run it.\n"
        "You can fix this in either of these ways:\n"
        "1. Install it into your current Python:\n"
        "   python -m pip install openpyxl\n"
        "2. Use the bundled Python runtime that already has openpyxl available."
        f"{bundled_hint}"
    )


@dataclass(frozen=True)
class ShiftDefinition:
    column_index: int
    label: str
    default_start: time
    default_end: time
    site_mode: str
    schedule_mode: str
    service_category: str
    admin_categories: tuple[str, ...]


@dataclass
class ShiftEntry:
    service_date: date
    start_time: time
    end_time: time
    site_mode: str
    schedule_mode: str
    service_category: str
    note: str | None = None
    source: str = ""
    admin_categories: tuple[str, ...] = ()
    display_order: int = 0


@dataclass
class GenerationOptions:
    template: Path
    master_url: str
    physician_name: str
    first_name: str
    last_name: str
    msp: str
    site: str
    schedule_name: str
    schedule_aliases: list[str]
    year: int
    month: int
    period: str
    clinical_admin_types: set[str]
    clinical_admin_note: str
    submission_date: date
    output: Path | None = None
    dry_run: bool = False


@dataclass
class GenerationResult:
    options: GenerationOptions
    aliases_used: list[str]
    matched_shifts: list[ShiftEntry]
    invoice_rows: list[ShiftEntry]
    added_admin_rows: list[ShiftEntry]
    skipped_admin_messages: list[str]
    output_path: Path | None = None


@dataclass
class ParsedFormData:
    values: dict[str, list[str]]
    uploaded_template_path: Path | None = None


TEAM_SHIFTS = [
    ShiftDefinition(
        i,
        f"Team {i - 1}",
        time(7, 0),
        time(17, 0),
        "On-Site",
        "Scheduled",
        "Patient Care",
        ("scheduled",),
    )
    for i in range(2, 11)
]
ADMIT_SHIFT = ShiftDefinition(
    11,
    "ADMIT",
    time(12, 0),
    time(17, 0),
    "On-Site",
    "Non-Scheduled",
    "Patient Care",
    ("admit",),
)
VIRTUAL_SHIFT = ShiftDefinition(
    12,
    "VIRTUAL",
    time(17, 0),
    time(21, 0),
    "Off-Site",
    "Non-Scheduled",
    "Virtual Shift",
    ("virtual",),
)
EVENING_SHIFT = ShiftDefinition(
    13,
    "Evening",
    time(17, 0),
    time(0, 0),
    "On-Site",
    "Scheduled",
    "Patient Care",
    ("evening",),
)
OVERNIGHT_SHIFT = ShiftDefinition(
    14,
    "Overnight",
    time(0, 0),
    time(7, 0),
    "On-Site",
    "Scheduled",
    "Patient Care",
    ("overnight",),
)


def build_parser() -> argparse.ArgumentParser:
    default_host = "0.0.0.0" if os.environ.get("PORT") else "127.0.0.1"
    default_port = int(os.environ.get("PORT", "8765"))
    parser = argparse.ArgumentParser(
        description="Generate a hospitalist invoice workbook from the Google Sheets master schedule."
    )
    parser.add_argument("--serve", action="store_true", help="Launch the local web form instead of running one-off CLI mode.")
    parser.add_argument("--host", default=default_host, help="Host for web mode.")
    parser.add_argument("--port", type=int, default=default_port, help="Port for web mode.")
    parser.add_argument("--template", type=Path, default=DEFAULT_TEMPLATE, help="Path to an existing invoice workbook.")
    parser.add_argument("--master-url", default=DEFAULT_MASTER_URL, help="Google Sheets share or export URL.")
    parser.add_argument("--physician-name", help="Physician legal name for the invoice header.")
    parser.add_argument("--first-name", help="First name for the invoice header.")
    parser.add_argument("--last-name", help="Last name for the invoice header.")
    parser.add_argument("--msp", help="MSP number.")
    parser.add_argument("--site", help="Site / facility code, for example PAH.")
    parser.add_argument("--schedule-name", help="Name as it appears on the master schedule.")
    parser.add_argument(
        "--schedule-alias",
        action="append",
        default=[],
        help="Additional master schedule aliases. Repeat for multiple aliases; use ; to separate multiple aliases in one value.",
    )
    parser.add_argument("--year", type=int, help="Invoice year, for example 2026.")
    parser.add_argument("--month", type=int, help="Invoice month number, for example 3 for March.")
    parser.add_argument(
        "--period",
        choices=["first", "second"],
        help="Use 'first' for days 1-15 or 'second' for days 16-end of month.",
    )
    parser.add_argument(
        "--clinical-admin-type",
        action="append",
        default=[],
        help="Shift types eligible for one 15-minute clinical admin add-on: scheduled, evening, overnight, admit, virtual.",
    )
    parser.add_argument(
        "--clinical-admin-note",
        default="doing billings",
        help="Optional note added to clinical admin rows.",
    )
    parser.add_argument("--submission-date", help="Submission date in YYYY-MM-DD format. Defaults to today.")
    parser.add_argument("--output", type=Path, help="Destination workbook path.")
    parser.add_argument("--dry-run", action="store_true", help="Print matched shifts without writing a workbook.")
    return parser


def prompt_if_missing(current: str | None, label: str, default: str | None = None) -> str:
    if current:
        return current
    suffix = f" [{default}]" if default else ""
    value = input(f"{label}{suffix}: ").strip()
    return value or (default or "")


def prompt_if_missing_int(current: int | None, label: str, default: int | None = None) -> int:
    if current is not None:
        return current
    while True:
        raw = prompt_if_missing(None, label, str(default) if default is not None else None)
        try:
            return int(raw)
        except ValueError:
            print(f"Please enter a number for {label.lower()}.")


def prompt_if_missing_choice(current: str | None, label: str, choices: Iterable[str], default: str) -> str:
    if current:
        return current
    choice_list = "/".join(choices)
    while True:
        value = prompt_if_missing(None, f"{label} ({choice_list})", default).lower()
        if value in choices:
            return value
        print(f"Please choose one of: {choice_list}.")


def prompt_for_admin_types(current_values: list[str]) -> set[str]:
    parsed = parse_admin_type_values(current_values)
    if parsed:
        return parsed
    selected: set[str] = set()
    for key, label in ADMIN_TYPE_LABELS.items():
        while True:
            answer = input(f"Add 15-minute clinical admin after {label}? [y/N]: ").strip().lower()
            if answer in {"", "n", "no"}:
                break
            if answer in {"y", "yes"}:
                selected.add(key)
                break
            print("Please answer y or n.")
    return selected


def normalize_text(value: str) -> str:
    return re.sub(r"\s+", " ", value.strip().lower())


def tokenize(value: str) -> list[str]:
    return re.findall(r"[a-z]+", normalize_text(value))


def alias_in_text(alias: str, text: str) -> bool:
    alias_tokens = tokenize(alias)
    text_tokens = tokenize(text)
    if not alias_tokens or not text_tokens or len(alias_tokens) > len(text_tokens):
        return False
    last_start = len(text_tokens) - len(alias_tokens) + 1
    for start in range(last_start):
        if text_tokens[start : start + len(alias_tokens)] == alias_tokens:
            return True
    return False


def split_aliases(raw_aliases: Iterable[str]) -> list[str]:
    aliases: list[str] = []
    for raw in raw_aliases:
        for part in re.split(r"[;|]", raw):
            cleaned = part.strip()
            if cleaned and cleaned not in aliases:
                aliases.append(cleaned)
    return aliases


def schedule_name_is_specific(value: str) -> bool:
    tokens = tokenize(value)
    normalized = normalize_text(value)
    return len(tokens) > 1 or "," in normalized or "." in normalized


def schedule_name_matches_person(
    schedule_name: str,
    first_name: str,
    last_name: str,
    physician_name: str,
) -> bool:
    schedule_name = schedule_name.strip()
    if not schedule_name:
        return False
    full_name = f"{first_name} {last_name}".strip()
    return any(
        [
            alias_in_text(schedule_name, physician_name),
            alias_in_text(schedule_name, full_name),
            alias_in_text(schedule_name, last_name),
            alias_in_text(last_name, schedule_name),
        ]
    )


def derive_default_aliases(first_name: str, last_name: str, physician_name: str, schedule_name: str | None) -> list[str]:
    aliases: list[str] = []
    schedule_name = (schedule_name or "").strip()
    candidates = [schedule_name]
    if not schedule_name or schedule_name_matches_person(schedule_name, first_name, last_name, physician_name):
        candidates.extend([physician_name, f"{first_name} {last_name}".strip()])
    if not schedule_name and last_name:
        candidates.extend(
            [
                last_name,
                f"{last_name}, {first_name[:1]}." if first_name else "",
                f"{last_name}, {first_name[:1]}" if first_name else "",
            ]
        )
    elif schedule_name_matches_person(schedule_name, first_name, last_name, physician_name) and not schedule_name_is_specific(schedule_name):
        candidates.extend(
            [
                last_name,
                f"{last_name}, {first_name[:1]}." if first_name else "",
                f"{last_name}, {first_name[:1]}" if first_name else "",
            ]
        )
    for candidate in candidates:
        candidate = candidate.strip()
        if candidate and candidate not in aliases:
            aliases.append(candidate)
    return aliases


def parse_admin_type_values(values: Iterable[str]) -> set[str]:
    result: set[str] = set()
    for raw in values:
        for part in re.split(r"[,;| ]+", raw.strip()):
            if not part:
                continue
            lowered = part.lower()
            if lowered == "all":
                return set(ADMIN_TYPE_LABELS)
            if lowered == "none":
                continue
            if lowered not in ADMIN_TYPE_LABELS:
                raise ValueError(
                    f"Unknown clinical admin shift type '{part}'. Use scheduled, evening, overnight, admit, or virtual."
                )
            result.add(lowered)
    return result


def share_url_to_csv_url(url: str) -> str:
    if "export?format=csv" in url:
        return url
    parsed = urlparse(url)
    match = re.search(r"/spreadsheets/d/([^/]+)", parsed.path)
    if not match:
        raise ValueError("Could not extract the Google Sheets document id from the master URL.")
    doc_id = match.group(1)
    gid = parse_qs(parsed.query).get("gid", ["0"])[0]
    return f"https://docs.google.com/spreadsheets/d/{doc_id}/export?format=csv&gid={gid}"


def fetch_master_rows(url: str) -> list[list[str]]:
    csv_url = share_url_to_csv_url(url)
    with urlopen(csv_url, timeout=30) as response:
        csv_text = response.read().decode("utf-8-sig")
    return list(csv.reader(io.StringIO(csv_text)))


def parse_month_heading(cell_value: str) -> tuple[int, int] | None:
    value = cell_value.strip()
    try:
        parsed = datetime.strptime(value, "%B %Y")
    except ValueError:
        return None
    return parsed.year, parsed.month


def parse_time_expression(part: str) -> tuple[time, time] | None:
    match = re.search(
        r"(\d{1,2})(?::(\d{2}))?\s*(am|pm)?\s*(?:-|to)\s*"
        r"(\d{1,2})(?::(\d{2}))?\s*(am|pm)?",
        part,
        re.IGNORECASE,
    )
    if not match:
        return None
    sh, sm, sampm, eh, em, eampm = match.groups()
    sampm = (sampm or eampm or "").lower()
    eampm = (eampm or sampm or "").lower()
    if not sampm or not eampm:
        return None
    return to_clock_time(int(sh), int(sm or 0), sampm), to_clock_time(int(eh), int(em or 0), eampm)


def to_clock_time(hour: int, minute: int, ampm: str) -> time:
    hour = hour % 12
    if ampm == "pm":
        hour += 12
    return time(hour, minute)


def extract_evening_interval(cell_value: str, aliases: list[str]) -> tuple[time, time] | None:
    if not any(alias_in_text(alias, cell_value) for alias in aliases):
        return None
    for segment in re.split(r"/|;", cell_value):
        if any(alias_in_text(alias, segment) for alias in aliases):
            parsed = parse_time_expression(segment)
            if parsed:
                return parsed
    parsed = parse_time_expression(cell_value)
    if parsed:
        return parsed
    return EVENING_SHIFT.default_start, EVENING_SHIFT.default_end


def matching_schedule_segments(cell_value: str, aliases: list[str]) -> list[str]:
    segments = [segment.strip() for segment in re.split(r"/|;", cell_value) if segment.strip()]
    matches = [segment for segment in segments if any(alias_in_text(alias, segment) for alias in aliases)]
    if matches:
        return matches
    return [cell_value] if any(alias_in_text(alias, cell_value) for alias in aliases) else []


def is_on_call_day_shift(cell_value: str, aliases: list[str]) -> bool:
    for segment in matching_schedule_segments(cell_value, aliases):
        for match in re.finditer(r"([A-Za-z][A-Za-z .,'-]*?)\s*\*C\b", segment, re.IGNORECASE):
            flagged_name = match.group(1).strip(" ,")
            if any(alias_in_text(alias, flagged_name) or alias_in_text(flagged_name, alias) for alias in aliases):
                return True
    return False


def date_in_requested_period(target: date, year: int, month: int, period: str) -> bool:
    if target.year != year or target.month != month:
        return False
    if period == "first":
        return 1 <= target.day <= 15
    return target.day >= 16


def row_matches_aliases(row_value: str, aliases: list[str]) -> bool:
    return bool(row_value and any(alias_in_text(alias, row_value) for alias in aliases))


def shift_note_for_definition(definition: ShiftDefinition) -> str | None:
    if definition in TEAM_SHIFTS:
        return definition.label
    if definition is ADMIT_SHIFT:
        return "Admit Pilot"
    if definition is VIRTUAL_SHIFT:
        return "Virtual"
    if definition is EVENING_SHIFT:
        return "Evening"
    return None


def overnight_note(source_date: date, includes_evening: bool = False) -> str:
    month_label = source_date.strftime("%b")
    base = f"Overnight ({month_label} {source_date.day})"
    if includes_evening:
        return f"Evening + {base}"
    return base


def parse_master_schedule(rows: list[list[str]], aliases: list[str], year: int, month: int, period: str) -> list[ShiftEntry]:
    entries: list[ShiftEntry] = []
    current_year_month: tuple[int, int] | None = None
    order_counter = 0

    for row in rows:
        if not row:
            continue
        heading = parse_month_heading(row[0]) if row[0] else None
        if heading:
            current_year_month = heading
            continue
        if current_year_month is None or len(row) < 15:
            continue
        if not row[1].strip().isdigit():
            continue

        service_date = date(current_year_month[0], current_year_month[1], int(row[1]))
        if not date_in_requested_period(service_date, year, month, period):
            continue

        evening_text = row[EVENING_SHIFT.column_index].strip()
        overnight_text = row[OVERNIGHT_SHIFT.column_index].strip()
        evening_interval = extract_evening_interval(evening_text, aliases) if evening_text else None
        overnight_match = row_matches_aliases(overnight_text, aliases)

        for definition in TEAM_SHIFTS + [ADMIT_SHIFT, VIRTUAL_SHIFT]:
            cell_value = row[definition.column_index].strip()
            if not row_matches_aliases(cell_value, aliases):
                continue
            start_time = definition.default_start
            end_time = definition.default_end
            if definition in TEAM_SHIFTS and is_on_call_day_shift(cell_value, aliases):
                start_time = ON_CALL_DAY_SHIFT_START
                end_time = ON_CALL_DAY_SHIFT_END
            entries.append(
                ShiftEntry(
                    service_date=service_date,
                    start_time=start_time,
                    end_time=end_time,
                    site_mode=definition.site_mode,
                    schedule_mode=definition.schedule_mode,
                    service_category=definition.service_category,
                    note=shift_note_for_definition(definition),
                    source=definition.label,
                    admin_categories=definition.admin_categories,
                    display_order=order_counter,
                )
            )
            order_counter += 1

        if overnight_match:
            start_time = evening_interval[0] if evening_interval else OVERNIGHT_SHIFT.default_start
            admin_categories = ("evening", "overnight") if evening_interval else OVERNIGHT_SHIFT.admin_categories
            entries.append(
                ShiftEntry(
                    service_date=service_date + timedelta(days=1),
                    start_time=start_time,
                    end_time=OVERNIGHT_SHIFT.default_end,
                    site_mode=OVERNIGHT_SHIFT.site_mode,
                    schedule_mode=OVERNIGHT_SHIFT.schedule_mode,
                    service_category=OVERNIGHT_SHIFT.service_category,
                    note=overnight_note(service_date, includes_evening=bool(evening_interval)),
                    source="Evening + Overnight" if evening_interval else "Overnight",
                    admin_categories=admin_categories,
                    display_order=order_counter,
                )
            )
            order_counter += 1
        elif evening_interval:
            entries.append(
                ShiftEntry(
                    service_date=service_date,
                    start_time=evening_interval[0],
                    end_time=evening_interval[1],
                    site_mode=EVENING_SHIFT.site_mode,
                    schedule_mode=EVENING_SHIFT.schedule_mode,
                    service_category=EVENING_SHIFT.service_category,
                    note=shift_note_for_definition(EVENING_SHIFT),
                    source="Evening",
                    admin_categories=EVENING_SHIFT.admin_categories,
                    display_order=order_counter,
                )
            )
            order_counter += 1

    return sorted(entries, key=lambda item: (item.service_date, item.start_time, item.display_order))


def actual_interval(entry: ShiftEntry) -> tuple[datetime, datetime]:
    start_dt = datetime.combine(entry.service_date, entry.start_time)
    end_dt = datetime.combine(entry.service_date, entry.end_time)
    if end_dt <= start_dt:
        end_dt += timedelta(days=1)
    return start_dt, end_dt


def overlaps_interval(candidate_start: datetime, candidate_end: datetime, occupied: Iterable[tuple[datetime, datetime]]) -> bool:
    for start_dt, end_dt in occupied:
        if candidate_start < end_dt and candidate_end > start_dt:
            return True
    return False


def clinical_admin_site_mode(shift: ShiftEntry) -> str:
    if "virtual" in shift.admin_categories:
        return "Off-Site"
    return "On-Site"


def build_invoice_rows(
    shifts: list[ShiftEntry],
    enabled_admin_types: set[str],
    admin_note: str,
) -> tuple[list[ShiftEntry], list[ShiftEntry], list[str]]:
    final_rows: list[ShiftEntry] = []
    added_admin_rows: list[ShiftEntry] = []
    skipped_messages: list[str] = []
    occupied_intervals = [actual_interval(shift) for shift in shifts]

    for shift in shifts:
        final_rows.append(shift)
        if not any(category in enabled_admin_types for category in shift.admin_categories):
            continue

        _, shift_end = actual_interval(shift)
        admin_start = shift_end
        admin_end = shift_end + timedelta(minutes=ADMIN_MINUTES)
        occupied_without_self = [
            interval
            for other_shift, interval in zip(shifts, occupied_intervals)
            if other_shift is not shift
        ]
        admin_occupied = occupied_without_self + [actual_interval(admin_row) for admin_row in added_admin_rows]

        if overlaps_interval(admin_start, admin_end, admin_occupied):
            skipped_messages.append(
                f"Skipped clinical admin after {shift.service_date.isoformat()} {shift.source} because it overlaps another billed interval."
            )
            continue

        admin_row = ShiftEntry(
            service_date=shift.service_date,
            start_time=admin_start.time(),
            end_time=admin_end.time(),
            site_mode=clinical_admin_site_mode(shift),
            schedule_mode="Non-Scheduled",
            service_category="Clinical Admin",
            note=admin_note,
            source=f"Clinical Admin after {shift.source}",
        )
        final_rows.append(admin_row)
        added_admin_rows.append(admin_row)

    return final_rows, added_admin_rows, skipped_messages


def clear_invoice_rows(invoice_sheet) -> None:
    for row in range(START_ROW, END_ROW + 1):
        for column in range(1, 8):
            invoice_sheet.cell(row=row, column=column).value = None


def write_invoice_rows(invoice_sheet, rows: list[ShiftEntry]) -> None:
    if len(rows) > (END_ROW - START_ROW + 1):
        raise ValueError("The template does not have enough blank invoice rows for all matched shifts.")
    for offset, entry in enumerate(rows):
        row = START_ROW + offset
        invoice_sheet.cell(row=row, column=1, value=entry.service_date)
        invoice_sheet.cell(row=row, column=2, value=entry.start_time)
        invoice_sheet.cell(row=row, column=3, value=entry.end_time)
        invoice_sheet.cell(row=row, column=4, value=entry.site_mode)
        invoice_sheet.cell(row=row, column=5, value=entry.schedule_mode)
        invoice_sheet.cell(row=row, column=6, value=entry.service_category)
        invoice_sheet.cell(row=row, column=7, value=entry.note)


def configure_invoice_header(
    workbook,
    physician_name: str,
    first_name: str,
    last_name: str,
    msp: str,
    site: str,
    submission_date: date,
) -> None:
    invoice = workbook["Invoice"]
    invoice["B3"] = physician_name
    invoice["E3"] = f"{first_name} {last_name}".strip()
    invoice["B5"] = str(msp)
    invoice["B11"] = site
    invoice["E5"] = submission_date


def parse_submission_date(raw_value: str | None) -> date:
    if not raw_value:
        return date.today()
    return datetime.strptime(raw_value, "%Y-%m-%d").date()


def default_output_path(
    output_arg: Path | None,
    physician_name: str,
    year: int,
    month: int,
    period: str,
    site: str,
) -> Path:
    if output_arg:
        return output_arg
    month_name = date(year, month, 1).strftime("%b %Y")
    period_label = "1st half" if period == "first" else "2nd half"
    safe_name = re.sub(r"[^A-Za-z0-9._-]+", "_", physician_name).strip("_")
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    return OUTPUT_DIR / f"{month_name} {period_label} - {safe_name} - {site} - {timestamp}.xlsx"


def recalculate_workbook_with_excel(workbook_path: Path) -> None:
    if os.name != "nt":
        return

    powershell = shutil.which("powershell") or shutil.which("pwsh")
    if not powershell:
        return

    escaped_path = str(workbook_path).replace("'", "''")
    script = (
        "$ErrorActionPreference='Stop';"
        f"$path='{escaped_path}';"
        "$wb=$null;$excel=$null;"
        "try {"
        "  $excel = New-Object -ComObject Excel.Application;"
        "  $excel.Visible = $false;"
        "  $excel.DisplayAlerts = $false;"
        "  $wb = $excel.Workbooks.Open($path);"
        "  $excel.CalculateFullRebuild();"
        "  Start-Sleep -Milliseconds 750;"
        "  $wb.Save();"
        "  $wb.Close($true);"
        "} finally {"
        "  if ($wb -ne $null) { [System.Runtime.InteropServices.Marshal]::ReleaseComObject($wb) | Out-Null }"
        "  if ($excel -ne $null) { $excel.Quit(); [System.Runtime.InteropServices.Marshal]::ReleaseComObject($excel) | Out-Null }"
        "}"
    )
    try:
        subprocess.run(
            [powershell, "-NoProfile", "-NonInteractive", "-Command", script],
            check=False,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            timeout=120,
        )
    except (OSError, subprocess.SubprocessError):
        return


def print_generation_summary(result: GenerationResult) -> None:
    print(f"Matched {len(result.invoice_rows)} invoice rows.")
    for entry in result.invoice_rows:
        note_suffix = f" | note: {entry.note}" if entry.note else ""
        print(
            f"  {entry.service_date.isoformat()} | {entry.start_time.strftime('%H:%M')}-{entry.end_time.strftime('%H:%M')} | "
            f"{entry.site_mode} | {entry.schedule_mode} | {entry.service_category} | {entry.source}{note_suffix}"
        )
    if result.skipped_admin_messages:
        print("\nClinical admin skipped:")
        for message in result.skipped_admin_messages:
            print(f"  - {message}")
    if result.output_path:
        print(f"\nSaved invoice workbook to:\n{result.output_path}")


def build_generation_options_from_args(args: argparse.Namespace) -> GenerationOptions:
    today = date.today()
    physician_name = prompt_if_missing(args.physician_name, "Physician legal name")
    first_name = prompt_if_missing(args.first_name, "First name")
    last_name = prompt_if_missing(args.last_name, "Last name")
    msp = prompt_if_missing(args.msp, "MSP #")
    site = prompt_if_missing(args.site, "Site / facility", "PAH").upper()
    year = prompt_if_missing_int(args.year, "Invoice year", today.year)
    month = prompt_if_missing_int(args.month, "Invoice month", today.month)
    period = prompt_if_missing_choice(args.period, "Period", ["first", "second"], "first")
    schedule_name = prompt_if_missing(args.schedule_name, "Name as shown on master schedule", last_name)

    aliases = split_aliases(args.schedule_alias)
    clinical_admin_types = prompt_for_admin_types(args.clinical_admin_type)

    return GenerationOptions(
        template=args.template,
        master_url=args.master_url,
        physician_name=physician_name,
        first_name=first_name,
        last_name=last_name,
        msp=msp,
        site=site,
        schedule_name=schedule_name,
        schedule_aliases=aliases,
        year=year,
        month=month,
        period=period,
        clinical_admin_types=clinical_admin_types,
        clinical_admin_note=args.clinical_admin_note,
        submission_date=parse_submission_date(args.submission_date),
        output=args.output,
        dry_run=args.dry_run,
    )


def generate_invoice(options: GenerationOptions) -> GenerationResult:
    ensure_openpyxl_available()

    aliases = split_aliases(options.schedule_aliases)
    for alias in derive_default_aliases(
        options.first_name,
        options.last_name,
        options.physician_name,
        options.schedule_name,
    ):
        if alias not in aliases:
            aliases.append(alias)

    master_rows = fetch_master_rows(options.master_url)
    matched_shifts = parse_master_schedule(master_rows, aliases, options.year, options.month, options.period)
    if not matched_shifts:
        raise ValueError(
            "No shifts matched the requested physician, schedule name/aliases, and date range. "
            "Try adjusting the schedule name or aliases."
        )

    invoice_rows, added_admin_rows, skipped_admin_messages = build_invoice_rows(
        matched_shifts,
        options.clinical_admin_types,
        options.clinical_admin_note,
    )

    result = GenerationResult(
        options=options,
        aliases_used=aliases,
        matched_shifts=matched_shifts,
        invoice_rows=invoice_rows,
        added_admin_rows=added_admin_rows,
        skipped_admin_messages=skipped_admin_messages,
    )

    if options.dry_run:
        return result

    if not options.template.exists():
        raise FileNotFoundError(f"Template workbook not found: {options.template}")

    output_path = default_output_path(
        options.output,
        options.physician_name,
        options.year,
        options.month,
        options.period,
        options.site,
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(options.template, output_path)

    workbook = load_workbook(output_path)
    configure_invoice_header(
        workbook=workbook,
        physician_name=options.physician_name,
        first_name=options.first_name,
        last_name=options.last_name,
        msp=options.msp,
        site=options.site,
        submission_date=options.submission_date,
    )
    clear_invoice_rows(workbook["Invoice"])
    write_invoice_rows(workbook["Invoice"], result.invoice_rows)
    workbook.calculation = CalcProperties(calcMode="auto", fullCalcOnLoad=True, forceFullCalc=True)
    workbook.save(output_path)
    recalculate_workbook_with_excel(output_path)
    result.output_path = output_path
    return result


def month_options(selected: int | None) -> str:
    options = []
    for month_number in range(1, 13):
        month_name = date(2000, month_number, 1).strftime("%B")
        selected_attr = " selected" if month_number == selected else ""
        options.append(f'<option value="{month_number}"{selected_attr}>{month_name}</option>')
    return "\n".join(options)


def render_form(
    values: dict[str, str] | None = None,
    errors: list[str] | None = None,
    result: GenerationResult | None = None,
) -> str:
    values = values or {}
    errors = errors or []
    checked_admin = set(values.get("clinical_admin_types", "").split(",")) if values.get("clinical_admin_types") else set()
    selected_month = int(values["month"]) if values.get("month", "").isdigit() else date.today().month
    period_value = values.get("period", "first")
    current_year = date.today().year

    def field(name: str, default: str = "") -> str:
        return html.escape(values.get(name, default))

    def checked(name: str) -> str:
        return " checked" if name in checked_admin else ""

    error_html = ""
    if errors:
        items = "".join(f"<li>{html.escape(message)}</li>" for message in errors)
        error_html = f'<div class="alert error"><strong>Please fix:</strong><ul>{items}</ul></div>'

    result_html = ""
    if result:
        rows = []
        for entry in result.invoice_rows:
            rows.append(
                "<tr>"
                f"<td>{entry.service_date.isoformat()}</td>"
                f"<td>{entry.start_time.strftime('%H:%M')}</td>"
                f"<td>{entry.end_time.strftime('%H:%M')}</td>"
                f"<td>{html.escape(entry.site_mode)}</td>"
                f"<td>{html.escape(entry.schedule_mode)}</td>"
                f"<td>{html.escape(entry.service_category)}</td>"
                f"<td>{html.escape(entry.note or '')}</td>"
                "</tr>"
            )
        skipped_html = ""
        if result.skipped_admin_messages:
            skipped_items = "".join(f"<li>{html.escape(message)}</li>" for message in result.skipped_admin_messages)
            skipped_html = f"<h3>Skipped Clinical Admin</h3><ul>{skipped_items}</ul>"
        output_html = ""
        if result.output_path:
            output_name = result.output_path.name
            output_html = (
                '<p class="success">'
                f'Workbook ready: <a href="/outputs/{quote(output_name)}">{html.escape(output_name)}</a>'
                "</p>"
            )
        result_html = (
            '<section class="panel results">'
            "<h2>Generated Invoice</h2>"
            f"{output_html}"
            f"<p>Matched {len(result.matched_shifts)} source shifts and added {len(result.added_admin_rows)} clinical admin rows.</p>"
            '<div class="table-wrap"><table><thead><tr>'
            "<th>Date</th><th>Start</th><th>End</th><th>Mode</th><th>Schedule</th><th>Category</th><th>Notes</th>"
            "</tr></thead><tbody>"
            f"{''.join(rows)}"
            "</tbody></table></div>"
            f"{skipped_html}"
            "</section>"
        )

    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Hospitalist Invoice Generator</title>
  <style>
    :root {{
      --bg: #f4efe4;
      --panel: #fffaf0;
      --ink: #1f2a2e;
      --muted: #5e6a6f;
      --accent: #0f766e;
      --accent-soft: #d7efe8;
      --border: #d4c8b4;
      --error: #8f2d2d;
      --error-bg: #f9e1e1;
      --shadow: 0 22px 50px rgba(55, 46, 32, 0.12);
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      font-family: "Segoe UI", Tahoma, sans-serif;
      background:
        radial-gradient(circle at top left, rgba(15, 118, 110, 0.14), transparent 32%),
        linear-gradient(180deg, #efe6d5 0%, var(--bg) 38%, #f8f5ee 100%);
      color: var(--ink);
    }}
    .shell {{
      max-width: 1180px;
      margin: 0 auto;
      padding: 28px 20px 48px;
    }}
    .hero {{
      padding: 18px 0 8px;
    }}
    .hero h1 {{
      margin: 0 0 10px;
      font-size: clamp(2rem, 4vw, 3.3rem);
      line-height: 1;
      letter-spacing: -0.03em;
    }}
    .hero p {{
      margin: 0;
      max-width: 760px;
      color: var(--muted);
      font-size: 1rem;
    }}
    .grid {{
      display: grid;
      grid-template-columns: minmax(320px, 440px) minmax(0, 1fr);
      gap: 22px;
      align-items: start;
      margin-top: 24px;
    }}
    .panel {{
      background: rgba(255, 250, 240, 0.94);
      border: 1px solid var(--border);
      border-radius: 24px;
      padding: 22px;
      box-shadow: var(--shadow);
      backdrop-filter: blur(6px);
    }}
    .panel h2 {{
      margin: 0 0 18px;
      font-size: 1.2rem;
    }}
    .section-title {{
      margin: 20px 0 8px;
      font-size: 0.9rem;
      text-transform: uppercase;
      letter-spacing: 0.08em;
      color: var(--muted);
    }}
    .two {{
      display: grid;
      grid-template-columns: 1fr 1fr;
      gap: 12px;
    }}
    label {{
      display: block;
      margin-bottom: 12px;
      font-weight: 600;
      font-size: 0.95rem;
    }}
    input, textarea, select {{
      width: 100%;
      margin-top: 6px;
      border: 1px solid var(--border);
      border-radius: 14px;
      padding: 12px 14px;
      font: inherit;
      color: var(--ink);
      background: #fffdf8;
    }}
    textarea {{
      min-height: 90px;
      resize: vertical;
    }}
    .checks {{
      display: grid;
      gap: 10px;
      grid-template-columns: 1fr 1fr;
      margin: 10px 0 18px;
    }}
    .check {{
      display: flex;
      align-items: center;
      gap: 10px;
      padding: 12px 14px;
      border: 1px solid var(--border);
      border-radius: 16px;
      background: #fffdf8;
      font-weight: 500;
    }}
    .check input {{
      width: auto;
      margin: 0;
      transform: scale(1.15);
    }}
    .actions {{
      display: flex;
      gap: 12px;
      flex-wrap: wrap;
      margin-top: 18px;
    }}
    button {{
      border: 0;
      border-radius: 999px;
      padding: 13px 18px;
      font: inherit;
      font-weight: 700;
      cursor: pointer;
      background: var(--accent);
      color: white;
      box-shadow: 0 10px 24px rgba(15, 118, 110, 0.22);
    }}
    .ghost {{
      background: var(--accent-soft);
      color: var(--accent);
      box-shadow: none;
    }}
    .alert {{
      border-radius: 16px;
      padding: 14px 16px;
      margin-bottom: 16px;
    }}
    .error {{
      border: 1px solid #e0aaaa;
      background: var(--error-bg);
      color: var(--error);
    }}
    .success {{
      padding: 12px 14px;
      border-radius: 14px;
      background: var(--accent-soft);
    }}
    .success a {{
      color: var(--accent);
      font-weight: 700;
    }}
    .table-wrap {{
      overflow: auto;
      border: 1px solid var(--border);
      border-radius: 18px;
      background: white;
    }}
    table {{
      width: 100%;
      border-collapse: collapse;
      min-width: 720px;
    }}
    th, td {{
      padding: 11px 12px;
      border-bottom: 1px solid #ece3d4;
      text-align: left;
      font-size: 0.93rem;
    }}
    th {{
      position: sticky;
      top: 0;
      background: #f8f2e7;
    }}
    .hint {{
      color: var(--muted);
      font-size: 0.9rem;
      margin-top: -4px;
      margin-bottom: 10px;
    }}
    @media (max-width: 900px) {{
      .grid {{
        grid-template-columns: 1fr;
      }}
      .checks, .two {{
        grid-template-columns: 1fr;
      }}
    }}
  </style>
</head>
<body>
  <div class="shell">
    <div class="hero">
      <h1>Hospitalist Invoice Generator</h1>
      <p>Fill in your billing details, choose which shift types should get a 15-minute clinical admin add-on, and generate the invoice workbook from the live master schedule.</p>
    </div>
    <div class="grid">
      <form class="panel" method="post" action="/generate" enctype="multipart/form-data">
        {error_html}
        <h2>Invoice Details</h2>
        <div class="two">
          <label>Physician Legal Name
            <input name="physician_name" value="{field('physician_name')}" required>
          </label>
          <label>MSP #
            <input name="msp" value="{field('msp')}" required>
          </label>
        </div>
        <div class="two">
          <label>First Name
            <input name="first_name" value="{field('first_name')}" required>
          </label>
          <label>Last Name
            <input name="last_name" value="{field('last_name')}" required>
          </label>
        </div>
        <div class="two">
          <label>Site / Facility
            <input name="site" value="{field('site', 'PAH')}" required>
          </label>
          <label>Name on Master Schedule
            <input name="schedule_name" value="{field('schedule_name')}" required>
            <span class="hint">Usually last name.</span>
          </label>
        </div>
        <label>Extra Schedule Aliases
          <textarea name="schedule_aliases" placeholder="Optional. Separate aliases with ;">{field('schedule_aliases')}</textarea>
        </label>
        <div class="section-title">Reporting Period</div>
        <div class="two">
          <label>Year
            <input name="year" type="number" value="{field('year', str(current_year))}" required>
          </label>
          <label>Month
            <select name="month">{month_options(selected_month)}</select>
          </label>
        </div>
        <div class="two">
          <label>Half of Month
            <select name="period">
              <option value="first"{' selected' if period_value == 'first' else ''}>1 to 15</option>
              <option value="second"{' selected' if period_value == 'second' else ''}>16 to month end</option>
            </select>
          </label>
          <label>Submission Date
            <input name="submission_date" type="date" value="{field('submission_date', date.today().isoformat())}">
          </label>
        </div>
        <div class="section-title">Clinical Admin</div>
        <p class="hint">These checkboxes apply once in general by shift type. If the 15-minute add-on would overlap another billed interval, it is skipped automatically.</p>
        <div class="checks">
          <label class="check"><input type="checkbox" name="clinical_admin_type" value="scheduled"{checked('scheduled')}>Scheduled / Team</label>
          <label class="check"><input type="checkbox" name="clinical_admin_type" value="evening"{checked('evening')}>Evening</label>
          <label class="check"><input type="checkbox" name="clinical_admin_type" value="overnight"{checked('overnight')}>Overnight</label>
          <label class="check"><input type="checkbox" name="clinical_admin_type" value="admit"{checked('admit')}>ADMIT</label>
          <label class="check"><input type="checkbox" name="clinical_admin_type" value="virtual"{checked('virtual')}>VIRTUAL</label>
        </div>
        <label>Clinical Admin Note
          <input name="clinical_admin_note" value="{field('clinical_admin_note', 'doing billings')}">
        </label>
        <div class="section-title">Sources</div>
        <label>Template Workbook Upload
          <input name="template_file" type="file" accept=".xlsx,.xlsm,.xltx,.xltm">
        </label>
        <p class="hint">A bundled template is used automatically. Upload a file here only if you want to override it for this one run.</p>
        <label>Template Workbook Path
          <input name="template" value="{field('template', str(DEFAULT_TEMPLATE))}">
        </label>
        <label>Master Schedule URL
          <input name="master_url" value="{field('master_url', DEFAULT_MASTER_URL)}" required>
        </label>
        <p class="hint">Leave the template path as-is unless you want to point to a different workbook on your own computer.</p>
        <div class="actions">
          <button type="submit">Generate Workbook</button>
          <button class="ghost" type="button" onclick="window.location='/'">Reset</button>
        </div>
      </form>
      {result_html or '<section class="panel"><h2>How It Works</h2><p>The app reads the live Google Sheets schedule, matches your name or aliases for the selected half-month, fills the invoice template, and saves the workbook into the local outputs folder.</p><p>Use aliases when the master sheet uses abbreviations or a slightly different spelling for your name.</p><p>If you host this on Railway, upload the template workbook in the form instead of using a local file path.</p></section>'}
    </div>
  </div>
</body>
</html>"""


def collect_form_value(data: dict[str, list[str]], key: str, default: str = "") -> str:
    values = data.get(key)
    if not values:
        return default
    return values[0].strip()


def build_options_from_form(data: dict[str, list[str]], uploaded_template_path: Path | None = None) -> GenerationOptions:
    physician_name = collect_form_value(data, "physician_name")
    first_name = collect_form_value(data, "first_name")
    last_name = collect_form_value(data, "last_name")
    msp = collect_form_value(data, "msp")
    site = collect_form_value(data, "site", "PAH").upper()
    schedule_name = collect_form_value(data, "schedule_name")
    aliases = split_aliases([collect_form_value(data, "schedule_aliases")])
    year_raw = collect_form_value(data, "year", str(date.today().year))
    month_raw = collect_form_value(data, "month", str(date.today().month))
    period = collect_form_value(data, "period", "first")
    admin_types = parse_admin_type_values(data.get("clinical_admin_type", []))

    missing = []
    for label, value in [
        ("Physician legal name", physician_name),
        ("First name", first_name),
        ("Last name", last_name),
        ("MSP #", msp),
        ("Site / facility", site),
        ("Name on master schedule", schedule_name),
    ]:
        if not value:
            missing.append(f"{label} is required.")
    if period not in {"first", "second"}:
        missing.append("Period must be first or second half of the month.")
    if missing:
        raise ValueError("\n".join(missing))

    try:
        year = int(year_raw)
        month = int(month_raw)
    except ValueError as exc:
        raise ValueError("Year and month must be valid numbers.") from exc
    if month < 1 or month > 12:
        raise ValueError("Month must be between 1 and 12.")

    template_raw = collect_form_value(data, "template", str(DEFAULT_TEMPLATE))
    template_path = uploaded_template_path or (Path(template_raw) if template_raw else None)
    if template_path is None and DEFAULT_TEMPLATE.exists():
        template_path = DEFAULT_TEMPLATE
    if template_path is None:
        raise ValueError("Please upload a template workbook or enter a local template path.")
    master_url = collect_form_value(data, "master_url", DEFAULT_MASTER_URL)
    submission_date = parse_submission_date(collect_form_value(data, "submission_date", date.today().isoformat()))
    clinical_admin_note = collect_form_value(data, "clinical_admin_note", "doing billings")

    return GenerationOptions(
        template=template_path,
        master_url=master_url,
        physician_name=physician_name,
        first_name=first_name,
        last_name=last_name,
        msp=msp,
        site=site,
        schedule_name=schedule_name,
        schedule_aliases=aliases,
        year=year,
        month=month,
        period=period,
        clinical_admin_types=admin_types,
        clinical_admin_note=clinical_admin_note,
        submission_date=submission_date,
    )


def form_values_from_request_data(data: dict[str, list[str]]) -> dict[str, str]:
    values = {
        "physician_name": collect_form_value(data, "physician_name"),
        "first_name": collect_form_value(data, "first_name"),
        "last_name": collect_form_value(data, "last_name"),
        "msp": collect_form_value(data, "msp"),
        "site": collect_form_value(data, "site", "PAH"),
        "schedule_name": collect_form_value(data, "schedule_name"),
        "schedule_aliases": collect_form_value(data, "schedule_aliases"),
        "year": collect_form_value(data, "year", str(date.today().year)),
        "month": collect_form_value(data, "month", str(date.today().month)),
        "period": collect_form_value(data, "period", "first"),
        "submission_date": collect_form_value(data, "submission_date", date.today().isoformat()),
        "clinical_admin_note": collect_form_value(data, "clinical_admin_note", "doing billings"),
        "template": collect_form_value(data, "template", str(DEFAULT_TEMPLATE)),
        "master_url": collect_form_value(data, "master_url", DEFAULT_MASTER_URL),
        "clinical_admin_types": ",".join(data.get("clinical_admin_type", [])),
    }
    return values


def save_uploaded_template(filename: str, content: bytes) -> Path | None:
    filename = Path(filename).name
    if not filename:
        return None

    if not content:
        return None

    upload_dir = OUTPUT_DIR / "_uploads"
    upload_dir.mkdir(parents=True, exist_ok=True)
    suffix = Path(filename).suffix or ".xlsx"
    target = upload_dir / f"template-{uuid4().hex}{suffix}"
    target.write_bytes(content)
    return target


def parse_form_payload(handler: BaseHTTPRequestHandler) -> ParsedFormData:
    content_type = handler.headers.get("Content-Type", "")
    if content_type.startswith("multipart/form-data"):
        content_length = int(handler.headers.get("Content-Length", "0"))
        body = handler.rfile.read(content_length)
        message = BytesParser(policy=policy.default).parsebytes(
            b"Content-Type: " + content_type.encode("utf-8") + b"\r\n"
            b"MIME-Version: 1.0\r\n\r\n" + body
        )
        values: dict[str, list[str]] = {}
        uploaded_template_path: Path | None = None
        for part in message.iter_parts():
            key = part.get_param("name", header="content-disposition")
            if not key:
                continue
            filename = part.get_filename()
            content = part.get_payload(decode=True) or b""
            if filename:
                if key == "template_file" and uploaded_template_path is None:
                    uploaded_template_path = save_uploaded_template(filename, content)
                continue
            charset = part.get_content_charset() or "utf-8"
            values.setdefault(key, []).append(content.decode(charset))
        return ParsedFormData(values=values, uploaded_template_path=uploaded_template_path)

    content_length = int(handler.headers.get("Content-Length", "0"))
    body = handler.rfile.read(content_length).decode("utf-8")
    return ParsedFormData(values=parse_qs(body, keep_blank_values=True))


class InvoiceRequestHandler(BaseHTTPRequestHandler):
    server_version = "HospitalistInvoiceHTTP/1.0"

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path == "/":
            self.respond_html(render_form())
            return
        if parsed.path.startswith("/outputs/"):
            self.serve_output_file(parsed.path.removeprefix("/outputs/"))
            return
        self.send_error(HTTPStatus.NOT_FOUND, "Page not found.")

    def do_POST(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path != "/generate":
            self.send_error(HTTPStatus.NOT_FOUND, "Page not found.")
            return

        payload = parse_form_payload(self)
        values = form_values_from_request_data(payload.values)

        try:
            options = build_options_from_form(payload.values, payload.uploaded_template_path)
            result = generate_invoice(options)
            self.respond_html(render_form(values=values, result=result))
        except Exception as exc:  # noqa: BLE001
            errors = str(exc).splitlines() or [str(exc)]
            self.respond_html(render_form(values=values, errors=errors), status=HTTPStatus.BAD_REQUEST)
        finally:
            if payload.uploaded_template_path and payload.uploaded_template_path.exists():
                payload.uploaded_template_path.unlink(missing_ok=True)

    def serve_output_file(self, file_name: str) -> None:
        safe_name = Path(unquote(file_name)).name
        target = (OUTPUT_DIR / safe_name).resolve()
        if OUTPUT_DIR.resolve() not in target.parents or not target.exists():
            self.send_error(HTTPStatus.NOT_FOUND, "File not found.")
            return

        data = target.read_bytes()
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.send_header("Content-Length", str(len(data)))
        self.send_header("Content-Disposition", f'attachment; filename="{target.name}"')
        self.end_headers()
        self.wfile.write(data)

    def respond_html(self, body: str, status: HTTPStatus = HTTPStatus.OK) -> None:
        encoded = body.encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(encoded)))
        self.end_headers()
        self.wfile.write(encoded)

    def log_message(self, format: str, *args) -> None:  # noqa: A003
        return


def run_web_server(host: str, port: int) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    server = ThreadingHTTPServer((host, port), InvoiceRequestHandler)
    print(f"Open http://{host}:{port} in your browser")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nStopping server.")
    finally:
        server.server_close()


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()
    ensure_openpyxl_available()

    if args.serve:
        run_web_server(args.host, args.port)
        return

    options = build_generation_options_from_args(args)
    result = generate_invoice(options)
    print_generation_summary(result)


if __name__ == "__main__":
    main()
